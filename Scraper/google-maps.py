import requests
import pandas as pd
from openpyxl import load_workbook

base_url = 'https://maps.googleapis.com/maps/api/geocode/json?'
API_KEY = 'AIzaSyCtUX7NH7muTrJ5k7oOekXZ2PUydKdT3PA'

def get_latandlon():
    ws = pd.read_excel(io = 'temp/dataFiltered.xlsx', sheet_name = 'Loc')
    lst = ws['Location']
    
    lstFinal = list(map(lambda x: help(x), lst))

    latCol = []
    lonCol = []

    for col in lstFinal:
        latCol.append(col[0])
        lonCol.append(col[1])

    ws = ws.transpose().append([latCol]).append([lonCol]).transpose()
    ws.columns = ['Location', 'Latitude', 'Longitude']

    writer = pd.ExcelWriter('dataComplete.xlsx', engine = 'openpyxl')
    book = load_workbook('dataFiltered.xlsx')
    writer.book = book
    ws.to_excel(writer, sheet_name = 'Location')
    writer.save()
    book2 = load_workbook('dataComplete.xlsx')
    book2.remove(book2['Loc'])
    book2.save('dataComplete.xlsx')

def help(address):
    response = requests.get(base_url, params = {'key': API_KEY, 
                                                'address': address}).json()
    response.keys()

    if response['status'] == 'OK':
        geometry = response['results'][0]['geometry']
        lat = geometry['location']['lat']
        lon = geometry['location']['lng']

    return [lat, lon]

get_latandlon()